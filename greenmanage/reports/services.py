import io
from openpyxl import Workbook
from transactions.models import Transaction


def generate_excel(user, start, end):
    wb = Workbook()

    data = Transaction.objects.filter(user=user, created_at__gte=start, created_at__lte=end)

    if not data.exists():
        ws = wb.active
        ws.title = "No Data"
        ws.append(["Номер", "Дата", "Описание", "Сумма", "Валюта", "Категория", "Тип транзакции"])
    else:
        ws = wb.create_sheet(title="Транзакции")
        ws.append(["Номер", "Дата", "Описание", "Сумма", "Валюта", "Категория", "Тип транзакции"])

        for index, transaction in enumerate(data, start=1):
            ws.append([
                index,
                transaction.created_at.strftime('%Y-%m-%d'),
                transaction.description,
                transaction.amount,
                transaction.currency.code,
                transaction.category.name,
                transaction.transaction_type.verbose_name
            ])

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    buffer = io.BytesIO()
    wb.save(buffer)

    buffer.seek(0)
    return buffer
